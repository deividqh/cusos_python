import os
os.system('cls')

# ----------------------------------------------------------------------------------------
# -------------------------------------- p a n d a s -------------------------------------
# ----------------------------------------------------------------------------------------

import pandas as XLS
# Leer un archivo Excel (pestaña por defecto)
pathArchivoXls="C:\\Users\\pc\\Desktop\\Personal\\ExcelDVD\\CONTROL-PADEL\\PAGOS - limpio.xlsm"

def xlsUsarPandas_Acceder():
    df = XLS.read_excel(pathArchivoXls)

    # Mostrar las primeras filas del DataFrame
    print(df.head())

    # Leer las primeras filas de una Hoja 
    df = XLS.read_excel(pathArchivoXls, sheet_name="Abril 24")
    print(df.head())


def xlsLeerTodasLasHojas():
    hojas = XLS.read_excel(pathArchivoXls, sheet_name=None)
    # Mostrar todas las hojas
    for nombre_hoja, df in hojas.items():
        print(f"Hoja: {nombre_hoja}")
        print(df.head())

# ---------Uso:
# xlsUsarPandas_Acceder() 
# xlsLeerTodasLasHojas()


# --------------------------------------------------------------------------------------------
# -------------------------------------- o p e n p y x l -------------------------------------
# --------------------------------------------------------------------------------------------

from openpyxl import load_workbook

def xlsOpenAcceso():
    # Cargar el archivo Excel
    wb = load_workbook(pathArchivoXls)

    # Seleccionar la hoja activa
    hoja = wb.active

    # Leer el valor de una celda
    valor = hoja["A1"].value
    print(f"Valor en A1: {valor}")

    # Leer varias celdas
    for fila in hoja.iter_rows(min_row=1, max_row=3, min_col=1, max_col=3, values_only=True):
        print(fila)

# ---------Uso:
# xlsOpenAcceso()



from openpyxl import Workbook

def xlsEscribirEnUnArchivoExcel():
    # Crear un nuevo archivo Excel
    wb = Workbook()

    # Seleccionar la hoja activa
    hoja = wb.active

    # Escribir datos en celdas
    hoja["A1"] = "Nombre"
    hoja["B1"] = "Edad"

    hoja["A2"] = "Juan"
    hoja["B2"] = 28

    # Guardar el archivo
    wb.save("salida.xlsx")

# ---------Uso:
# xlsEscribirEnUnArchivoExcel()

def ModificarUnArchivoExistente():
    # Cargar un archivo Excel existente
    wb = load_workbook("salida.xlsx")

    # Seleccionar la hoja activa
    hoja = wb.active

    # Modificar una celda
    hoja["A1"] = "Modificado"

    # Guardar los cambios
    wb.save("archivo_modificado.xlsx")

# ---------Uso:
# ModificarUnArchivoExistente()
# print(dir(XLS))
# print(dir(Workbook))
# print(dir(load_workbook))


from openpyxl import Workbook

def EscribirUnaFórmulaEnUnaCelda():
    # Crear un nuevo archivo Excel
    wb = Workbook()

    # Seleccionar la hoja activa
    hoja = wb.active

    # Escribir datos
    hoja["A1"] = 10
    hoja["A2"] = 20

    # Escribir una fórmula en una celda
    hoja["A3"] = "=SUM(A1:A2)"  # Fórmula para sumar los valores de A1 y A2

    # Guardar el archivo
    wb.save("archivo_con_formula.xlsx")

# ---------Uso:
EscribirUnaFórmulaEnUnaCelda()


from openpyxl import load_workbook
def LeerUnaFormula():
    # Cargar un archivo Excel existente
    wb = load_workbook("archivo_con_formula.xlsx", data_only=False)  # data_only=False para obtener las fórmulas
    hoja = wb.active
    # Leer la fórmula en la celda A3
    formula = hoja["A3"].value
    print(f"Fórmula en A3: {formula}")
# ---------Uso:

""" Para leer el resultado de una fórmula (en lugar de la fórmula misma), debes abrir el archivo Excel
 con la opción data_only=True. Sin embargo, ten en cuenta que el valor calculado solo estará disponible 
 si el archivo fue guardado previamente en Excel (o un programa que calcule las fórmulas).
 """

from openpyxl import load_workbook
def LeerResultadoDeFormula():
    # Cargar el archivo Excel existente
    wb = load_workbook("archivo_con_formula.xlsx", data_only=True)  # data_only=True para obtener el valor calculado
    hoja = wb.active

    # Leer el valor calculado en la celda A3
    valor = hoja["A3"].value
    print(f"Valor calculado en A3: {valor}")

# ---------Uso:

from openpyxl import Workbook
def ejemploEscribirVariasFormulas():
    # Crear un nuevo archivo Excel
    wb = Workbook()
    hoja = wb.active

    # Escribir datos
    hoja["A1"] = 5
    hoja["A2"] = 15
    hoja["B1"] = 10
    hoja["B2"] = 20

    # Escribir varias fórmulas
    hoja["C1"] = "=A1 + B1"  # Sumar A1 y B1
    hoja["C2"] = "=A2 * B2"  # Multiplicar A2 y B2
    hoja["C3"] = "=AVERAGE(A1:A2)"  # Promedio de A1 y A2
    hoja["C4"] = "=MAX(B1:B2)"  # Valor máximo de B1 y B2
    hoja["C5"] = "=IF(A1>B1, 'Mayor', 'Menor')"  # Condición

    # Guardar el archivo
    wb.save("archivo_varias_formulas.xlsx")

# ---------Uso:

# ------- Recalcular Formulas:
""" Una limitación de openpyxl es que no recalcula las fórmulas automáticamente dentro del archivo. 
El cálculo se realiza solo cuando el archivo es abierto en Excel o en otro programa que soporte
la recalculación de fórmulas.
"""
""" Para recalcular automáticamente las fórmulas, deberás abrir el archivo en Excel (u otro software compatible)
 y guardar los resultados.
"""

# --------------------------------- G R A F I C O S -------------------------------------------
# Gráfico de líneas: Usa LineChart.
# Gráfico de barras: Usa BarChart.
# Gráfico de pastel: Usa PieChart.
# Gráfico de dispersión: Usa ScatterChart.
# Gráficos combinados: Añadir múltiples series a un gráfico para combinarlos.
# ----------------------------------------------------------------------------

from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, PieChart, ScatterChart, Reference, Series

def grafico_lineas():
    wb = Workbook()
    hoja = wb.active

    datos = [
        ["Mes", "Ventas"],
        [1, 50],
        [2, 40],
        [3, 60],
        [4, 70],
        [5, 90],
    ]

    for fila in datos:
        hoja.append(fila)

    grafico = LineChart()
    valores = Reference(hoja, min_col=2, min_row=1, max_col=2, max_row=6)
    categorias = Reference(hoja, min_col=1, min_row=2, max_row=6)
    grafico.add_data(valores, titles_from_data=True)
    grafico.set_categories(categorias)
    grafico.title = "Ventas por Mes"
    grafico.y_axis.title = "Ventas"
    grafico.x_axis.title = "Mes"
    
    hoja.add_chart(grafico, "E1")
    wb.save("grafico_lineas.xlsx")

# ---------Uso:
# grafico_lineas()

# ---------------------------------------------------------
def grafico_barras():
    wb = Workbook()
    hoja = wb.active

    datos = [
        ["Producto", "Ventas"],
        ["Producto A", 50],
        ["Producto B", 40],
        ["Producto C", 60],
    ]

    for fila in datos:
        hoja.append(fila)

    grafico = BarChart()
    valores = Reference(hoja, min_col=2, min_row=1, max_col=2, max_row=4)
    categorias = Reference(hoja, min_col=1, min_row=2, max_row=4)
    grafico.add_data(valores, titles_from_data=True)
    grafico.set_categories(categorias)
    grafico.title = "Ventas por Producto"
    grafico.y_axis.title = "Ventas"
    grafico.x_axis.title = "Producto"
    
    hoja.add_chart(grafico, "E1")
    wb.save("grafico_barras.xlsx")

# ---------Uso:
grafico_barras()

# ---------------------------------------------------------
def grafico_pastel():
    wb = Workbook()
    hoja = wb.active

    datos = [
        ["Categoría", "Porcentaje"],
        ["A", 30],
        ["B", 20],
        ["C", 50],
    ]

    for fila in datos:
        hoja.append(fila)

    grafico = PieChart()
    valores = Reference(hoja, min_col=2, min_row=2, max_row=4)
    categorias = Reference(hoja, min_col=1, min_row=2, max_row=4)
    grafico.add_data(valores, titles_from_data=False)
    grafico.set_categories(categorias)
    grafico.title = "Distribución por Categoría"
    
    hoja.add_chart(grafico, "E1")
    wb.save("grafico_pastel.xlsx")

# ---------Uso:
# grafico_pastel()

# ---------------------------------------------------------
def grafico_dispersion():
    wb = Workbook()
    hoja = wb.active

    datos = [
        ["X", "Y"],
        [1, 2],
        [2, 3],
        [3, 5],
        [4, 7],
    ]

    for fila in datos:
        hoja.append(fila)

    grafico = ScatterChart()
    x_valores = Reference(hoja, min_col=1, min_row=2, max_row=5)
    y_valores = Reference(hoja, min_col=2, min_row=2, max_row=5)
    serie = Series(y_valores, xvalues=x_valores, title="Relación X vs Y")
    grafico.series.append(serie)
    grafico.title = "Gráfico de Dispersión"
    grafico.x_axis.title = "Valores X"
    grafico.y_axis.title = "Valores Y"
    
    hoja.add_chart(grafico, "E1")
    wb.save("grafico_dispersion.xlsx")

# ---------Uso:
# grafico_dispersion()

